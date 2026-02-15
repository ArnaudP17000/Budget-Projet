"""
Projet Manager - Business logic for project management.
"""
from typing import List, Optional, Dict
from database.db_manager import DatabaseManager
from database.models import Projet, InvestissementProjet, ContactSourcing, ProspectProjet
from utils.validators import validate_required_field, validate_montant, validate_date_range
from utils.constants import STATUTS_PROJET, TYPES_INVESTISSEMENT
from utils.formatters import parse_date


class ProjetManager:
    """Manages project business logic."""
    
    def __init__(self, db_manager: DatabaseManager):
        """Initialize with database manager."""
        self.db = db_manager
    
    def get_all_projets(self, statut: Optional[str] = None) -> List[Projet]:
        """Get all projects with optional status filter."""
        query = "SELECT * FROM projets WHERE 1=1"
        params = []
        
        if statut:
            query += " AND statut = ?"
            params.append(statut)
        
        query += " ORDER BY nom_projet"
        
        rows = self.db.execute_query(query, tuple(params))
        return [self._row_to_projet(row) for row in rows]
    
    def get_projet_by_id(self, projet_id: int) -> Optional[Projet]:
        """Get project by ID."""
        query = "SELECT * FROM projets WHERE id = ?"
        rows = self.db.execute_query(query, (projet_id,))
        if rows:
            return self._row_to_projet(rows[0])
        return None
    
    def create_projet(self, projet: Projet) -> tuple[bool, str, Optional[int]]:
        """Create new project."""
        # Validate required fields
        valid, msg = validate_required_field(projet.nom_projet, "Nom du projet")
        if not valid:
            return False, msg, None
        
        if projet.statut not in STATUTS_PROJET:
            return False, f"Statut invalide. Valeurs acceptées: {', '.join(STATUTS_PROJET)}", None
        
        # Validate date range
        if not validate_date_range(projet.date_debut, projet.date_fin_estimee):
            return False, "La date de fin doit être postérieure à la date de début"
        
        # Check if project name already exists
        existing = self.db.execute_query(
            "SELECT id FROM projets WHERE nom_projet = ?",
            (projet.nom_projet,)
        )
        if existing:
            return False, f"Un projet avec le nom '{projet.nom_projet}' existe déjà", None
        
        try:
            query = """
                INSERT INTO projets (nom_projet, fap_redigee, porteur_projet, service_demandeur,
                                   contacts_pris, sourcing, clients_contactes, pret_materiel_logiciel,
                                   date_debut, date_fin_estimee, date_mise_service,
                                   remarques_1, remarques_2, statut,
                                   investissement_licence, investissement_materiel, investissement_logiciel,
                                   cout_formation, frais_maintenance, technologies_utilisees)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            projet_id = self.db.execute_update(
                query,
                (projet.nom_projet, 1 if projet.fap_redigee else 0, projet.porteur_projet,
                 projet.service_demandeur, projet.contacts_pris, projet.sourcing,
                 projet.clients_contactes, projet.pret_materiel_logiciel,
                 projet.date_debut, projet.date_fin_estimee, projet.date_mise_service,
                 projet.remarques_1, projet.remarques_2, projet.statut,
                 projet.investissement_licence, projet.investissement_materiel, projet.investissement_logiciel,
                 projet.cout_formation, projet.frais_maintenance, projet.technologies_utilisees)
            )
            return True, "Projet créé avec succès", projet_id
        except Exception as e:
            return False, f"Erreur lors de la création: {str(e)}", None
    
    def update_projet(self, projet: Projet) -> tuple[bool, str]:
        """Update existing project."""
        if not projet.id:
            return False, "ID projet requis"
        
        # Validate required fields
        valid, msg = validate_required_field(projet.nom_projet, "Nom du projet")
        if not valid:
            return False, msg
        
        if projet.statut not in STATUTS_PROJET:
            return False, f"Statut invalide. Valeurs acceptées: {', '.join(STATUTS_PROJET)}"
        
        # Validate date range
        if not validate_date_range(projet.date_debut, projet.date_fin_estimee):
            return False, "La date de fin doit être postérieure à la date de début"
        
        # Check if project name already exists for another project
        existing = self.db.execute_query(
            "SELECT id FROM projets WHERE nom_projet = ? AND id != ?",
            (projet.nom_projet, projet.id)
        )
        if existing:
            return False, f"Un autre projet avec le nom '{projet.nom_projet}' existe déjà"
        
        try:
            query = """
                UPDATE projets
                SET nom_projet = ?, fap_redigee = ?, porteur_projet = ?, service_demandeur = ?,
                    contacts_pris = ?, sourcing = ?, clients_contactes = ?, pret_materiel_logiciel = ?,
                    date_debut = ?, date_fin_estimee = ?, date_mise_service = ?,
                    remarques_1 = ?, remarques_2 = ?, statut = ?,
                    investissement_licence = ?, investissement_materiel = ?, investissement_logiciel = ?,
                    cout_formation = ?, frais_maintenance = ?, technologies_utilisees = ?
                WHERE id = ?
            """
            self.db.execute_update(
                query,
                (projet.nom_projet, 1 if projet.fap_redigee else 0, projet.porteur_projet,
                 projet.service_demandeur, projet.contacts_pris, projet.sourcing,
                 projet.clients_contactes, projet.pret_materiel_logiciel,
                 projet.date_debut, projet.date_fin_estimee, projet.date_mise_service,
                 projet.remarques_1, projet.remarques_2, projet.statut,
                 projet.investissement_licence, projet.investissement_materiel, projet.investissement_logiciel,
                 projet.cout_formation, projet.frais_maintenance, projet.technologies_utilisees, projet.id)
            )
            return True, "Projet mis à jour avec succès"
        except Exception as e:
            return False, f"Erreur lors de la mise à jour: {str(e)}"
    
    def delete_projet(self, projet_id: int) -> tuple[bool, str]:
        """Delete project and all related data."""
        try:
            # Delete related data (cascade will handle this)
            query = "DELETE FROM projets WHERE id = ?"
            self.db.execute_update(query, (projet_id,))
            return True, "Projet supprimé avec succès"
        except Exception as e:
            return False, f"Erreur lors de la suppression: {str(e)}"
    
    # Investissement methods
    def get_investissements(self, projet_id: int) -> List[InvestissementProjet]:
        """Get all investments for a project."""
        query = "SELECT * FROM investissements_projets WHERE projet_id = ? ORDER BY type, description"
        rows = self.db.execute_query(query, (projet_id,))
        return [self._row_to_investissement(row) for row in rows]
    
    def add_investissement(self, investissement: InvestissementProjet) -> tuple[bool, str, Optional[int]]:
        """Add investment to project."""
        if not investissement.projet_id:
            return False, "ID projet requis", None
        
        valid, msg = validate_required_field(investissement.type, "Type")
        if not valid:
            return False, msg, None
        
        if investissement.type not in TYPES_INVESTISSEMENT:
            return False, f"Type invalide. Valeurs acceptées: {', '.join(TYPES_INVESTISSEMENT)}", None
        
        if not validate_montant(investissement.montant_estime):
            return False, "Montant estimé invalide", None
        
        try:
            query = """
                INSERT INTO investissements_projets (projet_id, type, description, montant_estime)
                VALUES (?, ?, ?, ?)
            """
            inv_id = self.db.execute_update(
                query,
                (investissement.projet_id, investissement.type,
                 investissement.description, investissement.montant_estime)
            )
            return True, "Investissement ajouté avec succès", inv_id
        except Exception as e:
            return False, f"Erreur lors de l'ajout: {str(e)}", None
    
    def update_investissement(self, investissement: InvestissementProjet) -> tuple[bool, str]:
        """Update investment."""
        if not investissement.id:
            return False, "ID investissement requis"
        
        if investissement.type not in TYPES_INVESTISSEMENT:
            return False, f"Type invalide. Valeurs acceptées: {', '.join(TYPES_INVESTISSEMENT)}"
        
        if not validate_montant(investissement.montant_estime):
            return False, "Montant estimé invalide"
        
        try:
            query = """
                UPDATE investissements_projets
                SET type = ?, description = ?, montant_estime = ?
                WHERE id = ?
            """
            self.db.execute_update(
                query,
                (investissement.type, investissement.description,
                 investissement.montant_estime, investissement.id)
            )
            return True, "Investissement mis à jour avec succès"
        except Exception as e:
            return False, f"Erreur lors de la mise à jour: {str(e)}"
    
    def delete_investissement(self, inv_id: int) -> tuple[bool, str]:
        """Delete investment."""
        try:
            query = "DELETE FROM investissements_projets WHERE id = ?"
            self.db.execute_update(query, (inv_id,))
            return True, "Investissement supprimé avec succès"
        except Exception as e:
            return False, f"Erreur lors de la suppression: {str(e)}"
    
    def get_total_investissements(self, projet_id: int) -> float:
        """Get total estimated investments for a project."""
        query = "SELECT SUM(montant_estime) as total FROM investissements_projets WHERE projet_id = ?"
        result = self.db.execute_query(query, (projet_id,))
        if result and result[0]['total']:
            return float(result[0]['total'])
        return 0.0
    
    # Contact sourcing methods
    def get_contacts_sourcing(self, projet_id: int) -> List[ContactSourcing]:
        """Get all sourcing contacts for a project."""
        query = "SELECT * FROM contacts_sourcing WHERE projet_id = ? ORDER BY nom, prenom"
        rows = self.db.execute_query(query, (projet_id,))
        return [self._row_to_contact_sourcing(row) for row in rows]
    
    def add_contact_sourcing(self, contact: ContactSourcing) -> tuple[bool, str, Optional[int]]:
        """Add sourcing contact to project."""
        if not contact.projet_id:
            return False, "ID projet requis", None
        
        valid, msg = validate_required_field(contact.nom, "Nom")
        if not valid:
            return False, msg, None
        
        valid, msg = validate_required_field(contact.prenom, "Prénom")
        if not valid:
            return False, msg, None
        
        try:
            query = """
                INSERT INTO contacts_sourcing (projet_id, nom, prenom, entreprise, telephone, email, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            """
            contact_id = self.db.execute_update(
                query,
                (contact.projet_id, contact.nom, contact.prenom, contact.entreprise,
                 contact.telephone, contact.email, contact.notes)
            )
            return True, "Contact sourcing ajouté avec succès", contact_id
        except Exception as e:
            return False, f"Erreur lors de l'ajout: {str(e)}", None
    
    def update_contact_sourcing(self, contact: ContactSourcing) -> tuple[bool, str]:
        """Update sourcing contact."""
        if not contact.id:
            return False, "ID contact requis"
        
        valid, msg = validate_required_field(contact.nom, "Nom")
        if not valid:
            return False, msg
        
        valid, msg = validate_required_field(contact.prenom, "Prénom")
        if not valid:
            return False, msg
        
        try:
            query = """
                UPDATE contacts_sourcing
                SET nom = ?, prenom = ?, entreprise = ?, telephone = ?, email = ?, notes = ?
                WHERE id = ?
            """
            self.db.execute_update(
                query,
                (contact.nom, contact.prenom, contact.entreprise,
                 contact.telephone, contact.email, contact.notes, contact.id)
            )
            return True, "Contact sourcing mis à jour avec succès"
        except Exception as e:
            return False, f"Erreur lors de la mise à jour: {str(e)}"
    
    def delete_contact_sourcing(self, contact_id: int) -> tuple[bool, str]:
        """Delete sourcing contact."""
        try:
            query = "DELETE FROM contacts_sourcing WHERE id = ?"
            self.db.execute_update(query, (contact_id,))
            return True, "Contact sourcing supprimé avec succès"
        except Exception as e:
            return False, f"Erreur lors de la suppression: {str(e)}"
    
    def _row_to_projet(self, row) -> Projet:
        """Convert database row to Projet object."""
        date_debut = parse_date(row['date_debut']) if row['date_debut'] else None
        date_fin_estimee = parse_date(row['date_fin_estimee']) if row['date_fin_estimee'] else None
        date_mise_service = parse_date(row['date_mise_service']) if row['date_mise_service'] else None
        
        return Projet(
            id=row['id'],
            nom_projet=row['nom_projet'],
            fap_redigee=bool(row['fap_redigee']),
            porteur_projet=row['porteur_projet'] or "",
            service_demandeur=row['service_demandeur'] or "",
            contacts_pris=row['contacts_pris'] or "",
            sourcing=row['sourcing'] or "",
            clients_contactes=row['clients_contactes'] or "",
            pret_materiel_logiciel=row['pret_materiel_logiciel'] or "",
            date_debut=date_debut,
            date_fin_estimee=date_fin_estimee,
            date_mise_service=date_mise_service,
            remarques_1=row['remarques_1'] or "",
            remarques_2=row['remarques_2'] or "",
            statut=row['statut'],
            investissement_licence=row.get('investissement_licence', 0) or 0,
            investissement_materiel=row.get('investissement_materiel', 0) or 0,
            investissement_logiciel=row.get('investissement_logiciel', 0) or 0,
            cout_formation=row.get('cout_formation', 0) or 0,
            frais_maintenance=row.get('frais_maintenance', 0) or 0,
            technologies_utilisees=row.get('technologies_utilisees', '') or ""
        )
    
    def _row_to_investissement(self, row) -> InvestissementProjet:
        """Convert database row to InvestissementProjet object."""
        return InvestissementProjet(
            id=row['id'],
            projet_id=row['projet_id'],
            type=row['type'],
            description=row['description'] or "",
            montant_estime=row['montant_estime']
        )
    
    def _row_to_contact_sourcing(self, row) -> ContactSourcing:
        """Convert database row to ContactSourcing object."""
        return ContactSourcing(
            id=row['id'],
            projet_id=row['projet_id'],
            nom=row['nom'],
            prenom=row['prenom'],
            entreprise=row['entreprise'] or "",
            telephone=row['telephone'] or "",
            email=row['email'] or "",
            notes=row['notes'] or ""
        )
    
    # Prospect management methods
    def add_prospect(self, projet_id, nom_prospect, description_offre="", 
                    investissement_licence=0, investissement_materiel=0,
                    investissement_logiciel=0, cout_formation=0, 
                    frais_maintenance=0, technologies="", notes=""):
        """Ajoute un prospect/fournisseur à un projet"""
        try:
            query = """
                INSERT INTO prospects_projets 
                (projet_id, nom_prospect, description_offre, investissement_licence,
                 investissement_materiel, investissement_logiciel, cout_formation,
                 frais_maintenance, technologies, notes)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
            self.db.execute_query(query, (projet_id, nom_prospect, description_offre,
                                          investissement_licence, investissement_materiel,
                                          investissement_logiciel, cout_formation,
                                          frais_maintenance, technologies, notes))
            return True, "Prospect ajouté avec succès"
        except Exception as e:
            return False, f"Erreur lors de l'ajout du prospect: {str(e)}"

    def update_prospect(self, prospect_id, **kwargs):
        """Met à jour un prospect"""
        try:
            fields = []
            values = []
            
            allowed_fields = ['nom_prospect', 'description_offre', 'investissement_licence',
                             'investissement_materiel', 'investissement_logiciel',
                             'cout_formation', 'frais_maintenance', 'technologies', 'notes']
            
            for field in allowed_fields:
                if field in kwargs:
                    fields.append(f"{field} = ?")
                    values.append(kwargs[field])
            
            if not fields:
                return False, "Aucun champ à mettre à jour"
            
            values.append(prospect_id)
            query = f"UPDATE prospects_projets SET {', '.join(fields)} WHERE id = ?"
            self.db.execute_query(query, tuple(values))
            
            return True, "Prospect mis à jour avec succès"
        except Exception as e:
            return False, f"Erreur lors de la mise à jour: {str(e)}"

    def delete_prospect(self, prospect_id):
        """Supprime un prospect"""
        try:
            query = "DELETE FROM prospects_projets WHERE id = ?"
            self.db.execute_query(query, (prospect_id,))
            return True, "Prospect supprimé avec succès"
        except Exception as e:
            return False, f"Erreur lors de la suppression: {str(e)}"

    def get_prospects_by_projet(self, projet_id):
        """Récupère tous les prospects d'un projet"""
        query = "SELECT * FROM prospects_projets WHERE projet_id = ? ORDER BY total_estime ASC"
        return self.db.fetch_all(query, (projet_id,))

    def export_projet_to_excel(self, projet_id, filepath):
        """Exporte un projet avec tous ses prospects en Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.chart import BarChart, Reference
            
            # Récupérer les données du projet
            query_projet = "SELECT * FROM projets WHERE id = ?"
            projet = self.db.fetch_one(query_projet, (projet_id,))
            
            if not projet:
                return False, "Projet introuvable"
            
            prospects = self.get_prospects_by_projet(projet_id)
            
            # Créer le workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Comparatif Prospects"
            
            # Styles
            header_fill = PatternFill(start_color="0d7377", end_color="0d7377", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Titre
            ws['A1'] = f"PROJET : {projet['nom_projet']}"
            ws['A1'].font = Font(bold=True, size=16, color="0d7377")
            ws.merge_cells('A1:H1')
            
            # Informations du projet
            row = 3
            ws[f'A{row}'] = "Porteur du projet:"
            ws[f'B{row}'] = projet['porteur_projet']
            row += 1
            ws[f'A{row}'] = "Service demandeur:"
            ws[f'B{row}'] = projet['service_demandeur']
            row += 1
            ws[f'A{row}'] = "Technologies:"
            ws[f'B{row}'] = projet.get('technologies_utilisees', '')
            row += 1
            ws[f'A{row}'] = "Statut:"
            ws[f'B{row}'] = projet['statut']
            
            # Comparatif des prospects
            row += 3
            ws[f'A{row}'] = "COMPARATIF DES PROSPECTS / FOURNISSEURS"
            ws[f'A{row}'].font = Font(bold=True, size=14, color="0d7377")
            ws.merge_cells(f'A{row}:H{row}')
            
            # En-têtes du tableau
            row += 2
            headers = ['Prospect', 'Licence', 'Matériel', 'Logiciel', 'Formation', 'Maintenance', 'TOTAL', 'Technologies']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            # Données des prospects
            for prospect in prospects:
                row += 1
                ws.cell(row=row, column=1, value=prospect['nom_prospect']).border = border
                ws.cell(row=row, column=2, value=prospect['investissement_licence']).border = border
                ws.cell(row=row, column=3, value=prospect['investissement_materiel']).border = border
                ws.cell(row=row, column=4, value=prospect['investissement_logiciel']).border = border
                ws.cell(row=row, column=5, value=prospect['cout_formation']).border = border
                ws.cell(row=row, column=6, value=prospect['frais_maintenance']).border = border
                ws.cell(row=row, column=7, value=prospect['total_estime']).border = border
                ws.cell(row=row, column=7).font = Font(bold=True)
                ws.cell(row=row, column=8, value=prospect['technologies']).border = border
                
                # Format monétaire
                for col in range(2, 8):
                    ws.cell(row=row, column=col).number_format = '#,##0.00 €'
            
            # Ajuster les largeurs de colonnes
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 30
            
            # Graphique de comparaison (si prospects)
            if len(prospects) > 0:
                chart_row = row + 3
                ws[f'A{chart_row}'] = "GRAPHIQUE DE COMPARAISON"
                ws[f'A{chart_row}'].font = Font(bold=True, size=12)
                
                chart = BarChart()
                chart.type = "col"
                chart.title = "Comparaison des coûts totaux"
                chart.y_axis.title = "Montant (€)"
                chart.x_axis.title = "Prospects"
                
                data = Reference(ws, min_col=7, min_row=row-len(prospects), max_row=row)
                cats = Reference(ws, min_col=1, min_row=row-len(prospects)+1, max_row=row)
                
                chart.add_data(data, titles_from_data=False)
                chart.set_categories(cats)
                chart.height = 10
                chart.width = 20
                
                ws.add_chart(chart, f'A{chart_row+2}')
            
            # Sauvegarder
            wb.save(filepath)
            
            return True, f"Export Excel créé avec succès : {filepath}"
            
        except ImportError:
            return False, "Module openpyxl non installé. Installez-le avec: pip install openpyxl"
        except Exception as e:
            return False, f"Erreur lors de l'export Excel: {str(e)}"
